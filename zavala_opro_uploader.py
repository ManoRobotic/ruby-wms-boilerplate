#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Zavala Production Order Uploader
Uploads production orders from opro.xlsx and oprod.xlsx to WMS API
"""

import requests
import time
import logging
from datetime import datetime
from openpyxl import load_workbook
import os
import json
import sys
from typing import Dict, List, Optional, Any

# Configure logging
log_filename = 'zavala_opro_uploader.log'
log_filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), log_filename) if os.path.dirname(os.path.abspath(__file__)) else log_filename
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filepath, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger('ZavalaOproUploader')

# Company configuration
COMPANY_NAME = "Rzavala"
COMPANY_TOKEN = "74bf5e0a6ae8813dfe80593ed84a7a9c"
WAREHOUSE_ID = "f7a1f77a-0802-49e3-871e-55bc917094f9"

# API configuration
API_BASE_URL = "https://wmsys.fly.dev"
API_ENDPOINT = "/api/production_orders/batch"
API_LAST_OPRO_ENDPOINT = "/api/production_orders/last_no_opro"
API_TIMEOUT = 90
MAX_RETRIES = 3
BATCH_SIZE = 25

# Configuration for file paths
if getattr(sys, 'frozen', False):
    application_path = sys._MEIPASS
    OPRO_XLSX_PATH = r"C:\ALPHAERP\Empresas\RZAVALA\opro.xlsx"
    OPROD_XLSX_PATH = r"C:\ALPHAERP\Empresas\RZAVALA\oprod.xlsx"
else:
    application_path = os.path.dirname(os.path.abspath(__file__))
    OPRO_XLSX_PATH = os.path.join(application_path, 'opro.xlsx')
    OPROD_XLSX_PATH = os.path.join(application_path, 'oprod.xlsx')

# State file to track last processed record
STATE_FILE = os.path.join(application_path, "zavala_opro_state.json")
LAST_MODIFIED_FILE = os.path.join(application_path, "zavala_opro_modified.json")


class ZavalaOproUploader:
    def __init__(self):
        self.session = requests.Session()
        self.state = self.load_state()
        self.last_modified_state = self.load_last_modified_state()
        self.first_run = not os.path.exists(STATE_FILE) and not os.path.exists(LAST_MODIFIED_FILE)
        self.sync_last_opro_from_api()

    def get_last_opro_from_api(self) -> int:
        """Get the last NO_OPRO from WMSys API for Zavala"""
        try:
            logger.info("Fetching last NO_OPRO from API...")
            response = self.session.get(
                API_BASE_URL + API_LAST_OPRO_ENDPOINT,
                headers={
                    'X-Company-Token': COMPANY_TOKEN
                },
                timeout=30
            )
            if response.status_code == 200:
                data = response.json()
                last_opro = data.get('last_no_opro', 0)
                logger.info(f"API returned last NO_OPRO: {last_opro}")
                return int(last_opro) if last_opro else 0
            else:
                logger.warning(f"API returned status {response.status_code}")
        except Exception as e:
            logger.warning(f"Could not fetch last NO_OPRO from API: {e}")
        return 0

    def sync_last_opro_from_api(self):
        """Sync the last processed NO_OPRO from API"""
        api_last_opro = self.get_last_opro_from_api()
        local_last_opro = self.state.get('last_processed_opro', 0)

        if api_last_opro > local_last_opro:
            logger.info(f"Syncing from API: {local_last_opro} -> {api_last_opro}")
            self.state['last_processed_opro'] = api_last_opro
            self.save_state()
        elif local_last_opro > api_last_opro:
            logger.info(f"Local state is ahead: {local_last_opro} > API: {api_last_opro}")

    def load_state(self) -> Dict:
        """Load the last processed state from file"""
        try:
            if os.path.exists(STATE_FILE):
                with open(STATE_FILE, 'r', encoding='utf-8') as f:
                    state = json.load(f)
                    if 'last_processed_opro' not in state:
                        state['last_processed_opro'] = 0
                    return state
        except Exception as e:
            logger.warning(f"Could not load state file: {e}")
        return {'last_processed_opro': 0}

    def save_state(self) -> bool:
        """Save the current state to file"""
        try:
            with open(STATE_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.state, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            logger.error(f"Error saving state: {e}")
            return False

    def load_last_modified_state(self) -> Dict:
        """Load the last modified timestamps from file"""
        try:
            if os.path.exists(LAST_MODIFIED_FILE):
                with open(LAST_MODIFIED_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.warning(f"Could not load last modified state file: {e}")
        return {}

    def save_last_modified_state(self) -> bool:
        """Save the current last modified timestamps to file"""
        try:
            with open(LAST_MODIFIED_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.last_modified_state, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            logger.error(f"Error saving last modified state: {e}")
            return False

    def get_file_last_modified(self, filepath: str) -> float:
        """Get the last modified timestamp of a file"""
        try:
            return os.path.getmtime(filepath)
        except Exception as e:
            logger.error(f"Error getting last modified time for {filepath}: {e}")
            return 0

    def has_file_changed(self, filepath: str) -> bool:
        """Check if a file has been modified since last check"""
        current_modified = self.get_file_last_modified(filepath)
        last_modified = self.last_modified_state.get(filepath, 0)

        if current_modified > last_modified:
            self.last_modified_state[filepath] = current_modified
            return True
        return False

    def clean_value(self, value: Any) -> str:
        """Clean and convert value to appropriate type"""
        if value is None or str(value).lower() in ['nan', 'none', '']:
            return ''
        if hasattr(value, 'strftime'):
            return value.isoformat()
        return str(value).strip()

    def extract_quantity(self, record: Dict) -> int:
        """Extract meaningful quantity from various fields"""
        try:
            ren_opro = self.clean_value(record.get('REN_OPRO', '0'))
            carga_opro = self.clean_value(record.get('CARGA_OPRO', '0'))
            can_op = self.clean_value(record.get('CAN_OP', '0'))

            for value in [ren_opro, carga_opro, can_op]:
                if value and value.lower() not in ['nan', 'none', '', '0']:
                    try:
                        qty = float(value)
                        if qty > 0:
                            return max(1, int(qty))
                    except:
                        continue

            return 1000
        except:
            return 1000

    def extract_year(self, record: Dict) -> str:
        """Extract year from date field"""
        try:
            fec_opro = self.clean_value(record.get('FEC_OPRO', ''))

            if fec_opro:
                if '-' in fec_opro:
                    return fec_opro.split('-')[0]
                elif '/' in fec_opro:
                    parts = fec_opro.split('/')
                    if len(parts) == 3:
                        return parts[2] if len(parts[2]) == 4 else ''
                elif len(fec_opro) >= 4:
                    if fec_opro[:4].isdigit():
                        return fec_opro[:4]

            return str(datetime.now().year)
        except:
            return str(datetime.now().year)

    def map_record_to_api(self, record: Dict) -> Optional[Dict]:
        """Map Excel record to API format"""
        try:
            cleaned = {k: self.clean_value(v) for k, v in record.items()}
            year = self.extract_year(cleaned)
            quantity = self.extract_quantity(cleaned)
            product_key = cleaned.get('CVE_PROP', '')
            no_opro = cleaned.get('NO_OPRO', '')

            if not no_opro:
                logger.warning("Skipping record: NO_OPRO is empty")
                return None

            if not product_key:
                logger.warning(f"Record with NO_OPRO {no_opro} has empty CVE_PROP")

            mapped = {
                "product_key": product_key,
                "quantity_requested": quantity,
                "warehouse_id": WAREHOUSE_ID,
                "priority": "medium",
                "no_opro": no_opro,
                "notes": cleaned.get('OBSERVA', ''),
                "lote_referencia": cleaned.get('LOTE', ''),
                "carga_copr": cleaned.get('CARGA_OPRO', ''),
                "ren_orp": cleaned.get('REN_OPRO', ''),
                "ano": year,
                "stat_opro": cleaned.get('STAT_OPRO', '')
            }

            mapped = {k: v for k, v in mapped.items() if v not in [None, 0] or k == 'notes'}

            logger.debug(f"Mapped record - NO_OPRO: {mapped.get('no_opro')}, "
                        f"Product: {mapped.get('product_key')}, "
                        f"Quantity: {mapped.get('quantity_requested')}")

            return mapped

        except Exception as e:
            logger.error(f"Error mapping record: {e}")
            return None

    def send_batch_to_api(self, batch_data: List[Dict]) -> Dict:
        """Send a batch of records to the API endpoint"""
        for attempt in range(MAX_RETRIES):
            try:
                logger.info(f"Sending batch of {len(batch_data)} records to API")

                payload = {
                    "company_name": COMPANY_NAME,
                    "production_orders": batch_data
                }

                response = self.session.post(
                    API_BASE_URL + API_ENDPOINT,
                    json=payload,
                    headers={
                        'Content-Type': 'application/json',
                        'X-Company-Token': COMPANY_TOKEN
                    },
                    timeout=API_TIMEOUT
                )

                logger.info(f"API Response Status: {response.status_code}")

                try:
                    response_content = response.json()
                    logger.debug(f"API Response Content: {json.dumps(response_content, indent=2)}")
                except:
                    logger.debug(f"API Response Text: {response.text}")
                    response_content = {}

                if response.status_code == 200:
                    try:
                        result = response_content
                        success_count = result.get('success_count', 0)
                        total_count = result.get('total_count', len(batch_data))
                        logger.info(f"API processed batch: {success_count}/{total_count} records successful")

                        existing_opros = []
                        for i, res in enumerate(result.get('results', [])):
                            if res.get('status') == 'error':
                                errors = res.get('errors', [])
                                logger.warning(f"Record {i} failed: {errors}")
                                error_str = str(errors).lower()
                                if "ya está en uso" in error_str or "already in use" in error_str:
                                    if i < len(batch_data):
                                        record = batch_data[i]
                                        opro = record.get('no_opro')
                                        if opro:
                                            existing_opros.append(opro)

                        return {
                            "success": True,
                            "data": result,
                            "existing_opros": existing_opros
                        }
                    except Exception as e:
                        logger.info(f"Batch sent successfully but error parsing response: {e}")
                        return {"success": True, "data": {}, "existing_opros": []}
                else:
                    logger.warning(f"API error {response.status_code}: {response.text}")
                    if attempt < MAX_RETRIES - 1:
                        time.sleep(2 ** attempt)

            except Exception as e:
                logger.error(f"Error sending batch (attempt {attempt + 1}): {e}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2 ** attempt)

        return {"success": False, "error": "Failed after retries", "existing_opros": []}

    def load_xlsx_data(self, filepath: str) -> List[Dict]:
        """Load data from XLSX file"""
        try:
            logger.info(f"Opening XLSX file: {filepath}")
            wb = load_workbook(filename=filepath, read_only=True)
            ws = wb.active

            headers = []
            records = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(row)]
                    continue

                record_dict = dict(zip(headers, row))
                records.append(record_dict)

            wb.close()
            logger.info(f"Loaded {len(records)} records from {filepath}")
            return records

        except Exception as e:
            logger.error(f"Error loading XLSX file {filepath}: {e}")
            return []

    def merge_opro_oprod(self, opro_records: List[Dict], oprod_records: List[Dict]) -> List[Dict]:
        """Merge opro and oprod records by NO_OPRO"""
        oprod_by_opro = {}
        for record in oprod_records:
            no_opro = str(record.get('NO_OPRO', ''))
            if no_opro:
                if no_opro not in oprod_by_opro:
                    oprod_by_opro[no_opro] = []
                oprod_by_opro[no_opro].append(record)

        merged_records = []
        for opro_record in opro_records:
            no_opro = str(opro_record.get('NO_OPRO', ''))
            if no_opro in oprod_by_opro:
                for oprod_record in oprod_by_opro[no_opro]:
                    merged = {**opro_record, **oprod_record}
                    merged_records.append(merged)
            else:
                merged_records.append(opro_record)

        logger.info(f"Merged {len(opro_records)} opro records with {len(oprod_records)} oprod records -> {len(merged_records)} total")
        return merged_records

    def process_xlsx_files(self) -> bool:
        """Process XLSX files and upload production orders"""
        try:
            logger.info("=" * 60)
            logger.info("PROCESSING ZAVALA PRODUCTION ORDERS")
            logger.info("=" * 60)

            opro_exists = os.path.exists(OPRO_XLSX_PATH)
            oprod_exists = os.path.exists(OPROD_XLSX_PATH)

            if not opro_exists:
                logger.error(f"File not found: {OPRO_XLSX_PATH}")
                return False

            opro_changed = self.first_run or self.has_file_changed(OPRO_XLSX_PATH)
            oprod_changed = self.first_run or (oprod_exists and self.has_file_changed(OPROD_XLSX_PATH))

            if not opro_changed and not oprod_changed:
                logger.info("No changes detected in XLSX files")
                self.save_last_modified_state()
                return True

            if self.first_run:
                logger.info("First run detected - will process all records")

            opro_records = self.load_xlsx_data(OPRO_XLSX_PATH)
            oprod_records = []
            if oprod_exists:
                oprod_records = self.load_xlsx_data(OPROD_XLSX_PATH)

            all_records = self.merge_opro_oprod(opro_records, oprod_records)

            processed_count = 0
            new_records = []
            current_last_opro = self.state.get('last_processed_opro', 0)
            logger.info(f"Starting process with known last NO_OPRO: {current_last_opro}")

            for record in all_records:
                no_opro = self.clean_value(record.get('NO_OPRO', ''))
                if not no_opro:
                    continue

                try:
                    current_opro_val = int(no_opro)
                    if current_opro_val > current_last_opro:
                        mapped_record = self.map_record_to_api(record)
                        if mapped_record:
                            new_records.append(mapped_record)
                            processed_count += 1
                except ValueError:
                    continue

            logger.info(f"Found {processed_count} new records based on NO_OPRO sequence")

            new_records.sort(key=lambda x: int(self.clean_value(x.get('no_opro', '0'))) if self.clean_value(x.get('no_opro', '0')).isdigit() else 0)

            if not new_records:
                logger.info("No new records to send based on NO_OPRO sequence")
                self.save_state()
                self.save_last_modified_state()
                self.first_run = False
                return True

            successful_sends = 0
            total_records = len(new_records)
            dynamic_last_opro = current_last_opro

            for i in range(0, len(new_records), BATCH_SIZE):
                batch = new_records[i:i + BATCH_SIZE]
                batch_num = i // BATCH_SIZE + 1

                try:
                    last_record_in_batch_opro = int(batch[-1]['no_opro'])
                    first_record_in_batch_opro = int(batch[0]['no_opro'])

                    if last_record_in_batch_opro <= dynamic_last_opro:
                        logger.info(f"Skipping batch {batch_num} (OPROs {first_record_in_batch_opro}-{last_record_in_batch_opro}) - already processed")
                        continue

                    if first_record_in_batch_opro <= dynamic_last_opro:
                        logger.info(f"Filtering batch {batch_num} - some records already processed")
                        batch = [r for r in batch if int(r['no_opro']) > dynamic_last_opro]
                        if not batch:
                            continue
                except (ValueError, IndexError):
                    pass

                logger.info(f"Processing batch {batch_num} ({len(batch)} records, OPROs {batch[0].get('no_opro')} - {batch[-1].get('no_opro')})")

                batch_result = self.send_batch_to_api(batch)

                if batch_result.get("success"):
                    result_data = batch_result.get("data", {})
                    success_count = result_data.get('success_count', len(batch))
                    successful_sends += success_count
                    logger.info(f"Batch {batch_num} sent: {success_count} records successful")

                    existing_opros = batch_result.get("existing_opros", [])
                    if existing_opros:
                        logger.info(f"Detected {len(existing_opros)} existing records in this batch.")
                        try:
                            existing_ints = [int(o) for o in existing_opros if str(o).isdigit()]
                            if existing_ints:
                                max_existing = max(existing_ints)
                                if max_existing > dynamic_last_opro:
                                    logger.info(f"Fast-forwarding state: {dynamic_last_opro} -> {max_existing}")
                                    dynamic_last_opro = max_existing

                                    if max_existing > self.state.get('last_processed_opro', 0):
                                        self.state['last_processed_opro'] = max_existing
                                        self.save_state()
                        except ValueError:
                            pass
                else:
                    logger.error(f"Batch {batch_num} failed: {batch_result.get('error')}")

            logger.info(f"Total records sent: {successful_sends}/{total_records}")

            if new_records:
                max_no_opro = max([int(r['no_opro']) for r in new_records if str(r['no_opro']).isdigit()], default=0)
                if max_no_opro > self.state.get('last_processed_opro', 0):
                    self.state['last_processed_opro'] = max_no_opro
                    self.save_state()
                    logger.info(f"Updated last_processed_opro to {max_no_opro}")

            self.save_last_modified_state()
            self.first_run = False

            return True

        except Exception as e:
            logger.error(f"Error processing XLSX files: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False


def main():
    logger.info("Starting Zavala Production Order Uploader")
    uploader = ZavalaOproUploader()
    success = uploader.process_xlsx_files()
    
    if success:
        logger.info("Zavala Production Order Uploader completed successfully")
        sys.exit(0)
    else:
        logger.error("Zavala Production Order Uploader failed")
        sys.exit(1)


if __name__ == "__main__":
    main()
