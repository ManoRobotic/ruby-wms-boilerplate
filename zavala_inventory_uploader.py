#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Zavala Inventory Code Uploader
Uploads inventory codes from remd.xlsx to WMS API

Company: Rzavala
Token: 74bf5e0a6ae8813dfe80593ed84a7a9c
Warehouse ID: f7a1f77a-0802-49e3-871e-55bc917094f9
"""

import requests
import time
import logging
from datetime import datetime
import os
import json
import sys
from typing import Dict, List, Optional, Any
from openpyxl import load_workbook

# Configure logging
log_filename = 'zavala_inventory_uploader.log'
log_filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), log_filename) if os.path.dirname(os.path.abspath(__file__)) else log_filename
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filepath, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger('ZavalaInventoryUploader')

# Company configuration
COMPANY_NAME = "Rzavala"
COMPANY_TOKEN = "74bf5e0a6ae8813dfe80593ed84a7a9c"
WAREHOUSE_ID = "f7a1f77a-0802-49e3-871e-55bc917094f9"

# API configuration
API_BASE_URL = "https://wmsys.fly.dev"
API_ENDPOINT = "/api/inventory_codes/batch"
API_TIMEOUT = 90
MAX_RETRIES = 3
BATCH_SIZE = 25

# Configuration for file paths
if getattr(sys, 'frozen', False):
    application_path = sys._MEIPASS
    XLSX_PATH = r"C:\ALPHAERP\Empresas\RZAVALA\remd.xlsx"
else:
    application_path = os.path.dirname(os.path.abspath(__file__))
    XLSX_PATH = os.path.join(application_path, 'remd.xlsx')

# State file to track last processed record
STATE_FILE = os.path.join(application_path, "zavala_inventory_state.json")
LAST_MODIFIED_FILE = os.path.join(application_path, "zavala_inventory_modified.json")


class ZavalaInventoryUploader:
    def __init__(self):
        self.session = requests.Session()
        self.state = self.load_state()
        self.last_modified_state = self.load_last_modified_state()
        self.first_run = not os.path.exists(STATE_FILE) and not os.path.exists(LAST_MODIFIED_FILE)
        self.sync_last_no_ordp_from_api()

    def get_last_no_ordp_from_api(self) -> int:
        """Get the last NO_ORDP from WMSys API for Zavala"""
        try:
            logger.info("Fetching last NO_ORDP from API...")
            response = self.session.get(
                API_BASE_URL + API_ENDPOINT.replace('/batch', '/sync_status'),
                headers={
                    'X-Company-Token': COMPANY_TOKEN
                },
                timeout=30
            )
            if response.status_code == 200:
                data = response.json()
                inventory_codes_count = data.get('inventory_codes_count', 0)
                logger.info(f"API returned inventory_codes_count: {inventory_codes_count}")
                # We use count as a reference, but actual last NO_ORDP should come from state
                return self.state.get('last_processed_no_ordp', 0)
            else:
                logger.warning(f"API returned status {response.status_code}")
        except Exception as e:
            logger.warning(f"Could not fetch from API: {e}")
        return 0

    def sync_last_no_ordp_from_api(self):
        """Sync the last processed NO_ORDP from API"""
        api_last = self.get_last_no_ordp_from_api()
        local_last = self.state.get('last_processed_no_ordp', 0)

        if local_last > api_last:
            logger.info(f"Local state is ahead: {local_last} > API count: {api_last}")
        else:
            logger.info(f"Syncing state - Local: {local_last}, API count: {api_last}")

    def load_state(self) -> Dict:
        """Load the last processed state from file"""
        try:
            if os.path.exists(STATE_FILE):
                with open(STATE_FILE, 'r', encoding='utf-8') as f:
                    state = json.load(f)
                    if 'last_processed_no_ordp' not in state:
                        state['last_processed_no_ordp'] = 0
                    return state
        except Exception as e:
            logger.warning(f"Could not load state file: {e}")
        return {'last_processed_no_ordp': 0}

    def save_state(self) -> bool:
        """Save the current state to file"""
        try:
            with open(STATE_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.state, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            logger.error(f"Error saving state: {e}")
            return False

    def load_last_modified_state(self) -> Dict:
        """Load the last modified timestamps from file"""
        try:
            if os.path.exists(LAST_MODIFIED_FILE):
                with open(LAST_MODIFIED_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.warning(f"Could not load last modified state file: {e}")
        return {}

    def save_last_modified_state(self) -> bool:
        """Save the current last modified timestamps to file"""
        try:
            with open(LAST_MODIFIED_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.last_modified_state, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            logger.error(f"Error saving last modified state: {e}")
            return False

    def get_file_last_modified(self, filepath: str) -> float:
        """Get the last modified timestamp of a file"""
        try:
            return os.path.getmtime(filepath)
        except Exception as e:
            logger.error(f"Error getting last modified time for {filepath}: {e}")
            return 0

    def has_file_changed(self, filepath: str) -> bool:
        """Check if a file has been modified since last check"""
        current_modified = self.get_file_last_modified(filepath)
        last_modified = self.last_modified_state.get(filepath, 0)

        if current_modified > last_modified:
            self.last_modified_state[filepath] = current_modified
            return True
        return False

    def clean_value(self, value: Any) -> str:
        """Clean and convert value to appropriate type"""
        if value is None or str(value).lower() in ['nan', 'none', '']:
            return ''
        if hasattr(value, 'strftime'):
            return value.isoformat()
        return str(value).strip()

    def map_record_to_api(self, record: Dict) -> Optional[Dict]:
        """Map Excel record to API format"""
        try:
            cleaned = {k: self.clean_value(v) for k, v in record.items()}

            no_ordp = cleaned.get('NO_ORDP', '')
            cve_prod = cleaned.get('CVE_PROD', '')
            cve_copr = cleaned.get('CVE_COPR', cleaned.get('CVE_PROD', ''))
            can_copr = cleaned.get('CAN_PROD', cleaned.get('CANT_PROD', '0'))
            lote = cleaned.get('LOTE', '')
            fech_cto = cleaned.get('FECH_ORDP', '')
            tip_copr = 1  # Default to active

            # Validate required fields
            if not no_ordp:
                logger.warning("Skipping record: NO_ORDP is empty")
                return None

            if not cve_prod:
                logger.warning(f"Record with NO_ORDP {no_ordp} has empty CVE_PROD")
                return None

            # Parse date if available
            parsed_date = None
            if fech_cto:
                try:
                    if 'T' in str(fech_cto):
                        parsed_date = fech_cto.split('T')[0]
                    else:
                        parsed_date = fech_cto
                except Exception:
                    parsed_date = datetime.now().strftime('%Y-%m-%d')
            else:
                parsed_date = datetime.now().strftime('%Y-%m-%d')

            # Parse quantity
            try:
                quantity = float(can_copr) if can_copr else 0
                if quantity <= 0:
                    quantity = 1
            except ValueError:
                quantity = 1

            mapped = {
                "no_ordp": no_ordp,
                "cve_prod": cve_prod,
                "cve_copr": cve_copr,
                "can_copr": quantity,
                "lote": lote,
                "fech_cto": parsed_date,
                "tip_copr": tip_copr
            }

            # Remove empty fields
            mapped = {k: v for k, v in mapped.items() if v not in [None, '', 0] or k in ['tip_copr', 'can_copr']}

            logger.debug(f"Mapped record - NO_ORDP: {mapped.get('no_ordp')}, "
                        f"Product: {mapped.get('cve_prod')}, "
                        f"Quantity: {mapped.get('can_copr')}")

            return mapped

        except Exception as e:
            logger.error(f"Error mapping record: {e}")
            return None

    def send_batch_to_api(self, batch_data: List[Dict]) -> Dict:
        """Send a batch of records to the API endpoint"""
        for attempt in range(MAX_RETRIES):
            try:
                logger.info(f"Sending batch of {len(batch_data)} records to API")

                payload = {
                    "company_name": COMPANY_NAME,
                    "inventory_codes": batch_data
                }

                response = self.session.post(
                    API_BASE_URL + API_ENDPOINT,
                    json=payload,
                    headers={
                        'Content-Type': 'application/json',
                        'X-Company-Token': COMPANY_TOKEN
                    },
                    timeout=API_TIMEOUT
                )

                logger.info(f"API Response Status: {response.status_code}")

                try:
                    response_content = response.json()
                    logger.debug(f"API Response Content: {json.dumps(response_content, indent=2)}")
                except:
                    logger.debug(f"API Response Text: {response.text}")
                    response_content = {}

                if response.status_code == 200:
                    try:
                        result = response_content
                        success_count = result.get('success_count', 0)
                        total_count = result.get('total_count', len(batch_data))
                        logger.info(f"API processed batch: {success_count}/{total_count} records successful")

                        return {
                            "success": True,
                            "data": result
                        }
                    except Exception as e:
                        logger.info(f"Batch sent successfully but error parsing response: {e}")
                        return {"success": True, "data": {}}
                else:
                    logger.warning(f"API error {response.status_code}: {response.text}")
                    if attempt < MAX_RETRIES - 1:
                        time.sleep(2 ** attempt)

            except Exception as e:
                logger.error(f"Error sending batch (attempt {attempt + 1}): {e}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2 ** attempt)

        return {"success": False, "error": "Failed after retries"}

    def process_xlsx_file(self) -> bool:
        """Process XLSX file and upload inventory codes"""
        try:
            logger.info("=" * 60)
            logger.info("PROCESSING ZAVALA INVENTORY CODES FROM REMD.XLSX")
            logger.info("=" * 60)

            if not os.path.exists(XLSX_PATH):
                logger.error(f"File not found: {XLSX_PATH}")
                return False

            if not self.first_run and not self.has_file_changed(XLSX_PATH):
                logger.info("No changes detected in XLSX file")
                self.save_last_modified_state()
                return True

            if self.first_run:
                logger.info("First run detected - will process all records")

            logger.info(f"Opening XLSX file: {XLSX_PATH}")
            wb = load_workbook(filename=XLSX_PATH, read_only=True)
            ws = wb.active

            headers = []
            all_records = []
            processed_count = 0
            new_records_count = 0

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(row)]
                    logger.info(f"Headers: {headers}")
                    continue

                record_dict = dict(zip(headers, row))
                no_ordp = self.clean_value(record_dict.get('NO_ORDP', ''))

                if not no_ordp:
                    continue

                new_records_count += 1
                mapped_record = self.map_record_to_api(record_dict)
                if mapped_record:
                    all_records.append(mapped_record)
                    processed_count += 1

                    if processed_count % 100 == 0:
                        logger.info(f"Processed {processed_count} records so far...")

            wb.close()

            logger.info(f"Found {new_records_count} records, prepared {len(all_records)} valid records for sending")

            all_records.sort(key=lambda x: str(x.get('no_ordp', '0')))

            if not all_records:
                logger.info("No valid records to send")
                self.save_state()
                self.save_last_modified_state()
                self.first_run = False
                return True

            successful_sends = 0
            total_records = len(all_records)

            for i in range(0, len(all_records), BATCH_SIZE):
                batch = all_records[i:i + BATCH_SIZE]
                batch_num = i // BATCH_SIZE + 1

                logger.info(f"Processing batch {batch_num} ({len(batch)} records)")

                batch_result = self.send_batch_to_api(batch)

                if batch_result.get("success"):
                    result_data = batch_result.get("data", {})
                    success_count = result_data.get('success_count', len(batch))
                    successful_sends += success_count
                    logger.info(f"Batch {batch_num} sent: {success_count} records successful")
                else:
                    logger.error(f"Batch {batch_num} failed: {batch_result.get('error')}")

            logger.info(f"Total records sent: {successful_sends}/{total_records}")

            if all_records:
                max_no_ordp = max([int(r['no_ordp']) for r in all_records if str(r['no_ordp']).isdigit()], default=0)
                if max_no_ordp > self.state.get('last_processed_no_ordp', 0):
                    self.state['last_processed_no_ordp'] = max_no_ordp
                    self.save_state()
                    logger.info(f"Updated last_processed_no_ordp to {max_no_ordp}")

            self.save_last_modified_state()
            self.first_run = False

            return True

        except Exception as e:
            logger.error(f"Error processing XLSX file: {e}")
            return False


def main():
    logger.info("Starting Zavala Inventory Uploader")
    uploader = ZavalaInventoryUploader()
    success = uploader.process_xlsx_file()
    
    if success:
        logger.info("Zavala Inventory Uploader completed successfully")
        sys.exit(0)
    else:
        logger.error("Zavala Inventory Uploader failed")
        sys.exit(1)


if __name__ == "__main__":
    main()
